from flask import Blueprint, request, jsonify, current_app as app
from app.dao.gestionar_personas.paciente.PacienteDao import PacienteDao
from flask import send_file
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
import io


pacienteapi = Blueprint('pacienteapi', __name__)


# ============================================
# GENERAR PDF DE PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/pdf', methods=['GET'])
def generarPDF(pac_id):
    """Genera un PDF con la ficha del paciente"""
    pacientedao = PacienteDao()
    
    try:
        paciente = pacientedao.getPacienteById(pac_id)
        
        if not paciente:
            return jsonify({'success': False, 'error': 'Paciente no encontrado'}), 404
        
        # Crear PDF en memoria
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        
        # Título
        p.setFont("Helvetica-Bold", 16)
        p.drawString(50, height - 50, "Ficha de Paciente")
        
        # Datos del paciente
        p.setFont("Helvetica", 12)
        y = height - 100
        
        datos = [
            f"Historia Clínica: {paciente.get('historia_clinica', 'N/A')}",
            f"Nombre: {paciente.get('nombre', '')} {paciente.get('apellido', '')}",
            f"Cédula: {paciente.get('cedula', 'N/A')}",
            f"Fecha Nacimiento: {paciente.get('fecha_nacimiento', 'N/A')}",
            f"Edad: {paciente.get('edad', 'N/A')} años",
            f"Es menor de edad: {'Sí' if paciente.get('es_menor') else 'No'}",
            f"Género: {paciente.get('genero', 'N/A')}",
            f"Estado Civil: {paciente.get('estado_civil', 'N/A')}",
            f"Teléfono: {paciente.get('telefono', 'N/A')}",
            f"Correo: {paciente.get('correo', 'N/A')}",
            f"Domicilio: {paciente.get('domicilio', 'N/A')}",
            f"Ciudad: {paciente.get('ciudad', 'N/A')}",
            f"Nivel Instrucción: {paciente.get('nivel_instruccion', 'N/A')}",
            f"Profesión: {paciente.get('profesion', 'N/A')}",
        ]
        
        for dato in datos:
            p.drawString(50, y, dato)
            y -= 20
        
        # Datos del tutor si es menor
        if paciente.get('es_menor') and (paciente.get('nom_madre') or paciente.get('nom_padre')):
            y -= 20
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, "Datos del Tutor:")
            y -= 20
            p.setFont("Helvetica", 12)
            
            if paciente.get('nom_madre'):
                p.drawString(50, y, f"Madre: {paciente.get('nom_madre', 'N/A')} - Tel: {paciente.get('tel_madre', 'N/A')}")
                y -= 20
            
            if paciente.get('nom_padre'):
                p.drawString(50, y, f"Padre: {paciente.get('nom_padre', 'N/A')} - Tel: {paciente.get('tel_padre', 'N/A')}")
                y -= 20
            
            if paciente.get('colegio'):
                p.drawString(50, y, f"Colegio: {paciente.get('colegio', 'N/A')} - Tel: {paciente.get('tel_colegio', 'N/A')}")
        
        if paciente.get('observaciones'):
            y -= 30
            p.setFont("Helvetica-Bold", 12)
            p.drawString(50, y, "Observaciones:")
            y -= 20
            p.setFont("Helvetica", 10)
            p.drawString(50, y, paciente.get('observaciones', ''))
        
        p.showPage()
        p.save()
        
        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name=f"paciente_{pac_id}.pdf", mimetype='application/pdf')
        
    except Exception as e:
        app.logger.error(f"Error al generar PDF: {str(e)}")
        return jsonify({'success': False, 'error': 'Error al generar PDF'}), 500


# ============================================
# GENERAR EXCEL DE PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/excel', methods=['GET'])
def generarExcel(pac_id):
    """Genera un archivo Excel con los datos del paciente"""
    pacientedao = PacienteDao()
    
    try:
        paciente = pacientedao.getPacienteById(pac_id)
        
        if not paciente:
            return jsonify({'success': False, 'error': 'Paciente no encontrado'}), 404
        
        # Crear Excel en memoria
        wb = Workbook()
        ws = wb.active
        ws.title = "Datos del Paciente"
        
        # Encabezados
        ws['A1'] = "Campo"
        ws['B1'] = "Valor"
        
        # Datos básicos
        datos = [
            ["Historia Clínica", paciente.get('historia_clinica', 'N/A')],
            ["Nombre", paciente.get('nombre', '')],
            ["Apellido", paciente.get('apellido', '')],
            ["Cédula", paciente.get('cedula', 'N/A')],
            ["Fecha Nacimiento", paciente.get('fecha_nacimiento', 'N/A')],
            ["Edad", f"{paciente.get('edad', 'N/A')} años"],
            ["Es menor de edad", 'Sí' if paciente.get('es_menor') else 'No'],
            ["Género", paciente.get('genero', 'N/A')],
            ["Estado Civil", paciente.get('estado_civil', 'N/A')],
            ["Teléfono", paciente.get('telefono', 'N/A')],
            ["Correo", paciente.get('correo', 'N/A')],
            ["Domicilio", paciente.get('domicilio', 'N/A')],
            ["Ciudad", paciente.get('ciudad', 'N/A')],
            ["Ciudad Nacimiento", paciente.get('ciudad_nacimiento', 'N/A')],
            ["Nivel Instrucción", paciente.get('nivel_instruccion', 'N/A')],
            ["Profesión", paciente.get('profesion', 'N/A')],
        ]
        
        # Agregar datos del tutor si es menor
        if paciente.get('es_menor'):
            datos.extend([
                ["--- Datos del Tutor ---", ""],
                ["Nombre Madre", paciente.get('nom_madre', 'N/A')],
                ["Teléfono Madre", paciente.get('tel_madre', 'N/A')],
                ["Nombre Padre", paciente.get('nom_padre', 'N/A')],
                ["Teléfono Padre", paciente.get('tel_padre', 'N/A')],
                ["Educación", paciente.get('educacion', 'N/A')],
                ["Colegio", paciente.get('colegio', 'N/A')],
                ["Teléfono Colegio", paciente.get('tel_colegio', 'N/A')],
            ])
        
        if paciente.get('observaciones'):
            datos.append(["Observaciones", paciente.get('observaciones', '')])
        
        for idx, (campo, valor) in enumerate(datos, start=2):
            ws[f'A{idx}'] = campo
            ws[f'B{idx}'] = valor
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(buffer, as_attachment=True, download_name=f"paciente_{pac_id}.xlsx", 
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
    except Exception as e:
        app.logger.error(f"Error al generar Excel: {str(e)}")
        return jsonify({'success': False, 'error': 'Error al generar Excel'}), 500


# ============================================
# OBTENER TODOS LOS PACIENTES
# ============================================
@pacienteapi.route('/pacientes', methods=['GET'])
def getPacientes():
    """Obtiene la lista completa de pacientes"""
    pacientedao = PacienteDao()
    
    try:
        pacientes = pacientedao.getPacientes()
        
        return jsonify({
            'success': True,
            'data': pacientes,
            'error': None
        }), 200
    
    except Exception as e:
        app.logger.error(f"Error al obtener todos los pacientes: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER SOLO PACIENTES MENORES
# ============================================
@pacienteapi.route('/pacientes/menores', methods=['GET'])
def getPacientesMenores():
    """Obtiene solo los pacientes menores de edad (calculado automáticamente)"""
    pacientedao = PacienteDao()
    
    try:
        menores = pacientedao.getPacientesMenores()
        
        return jsonify({
            'success': True,
            'data': menores,
            'error': None
        }), 200
    
    except Exception as e:
        app.logger.error(f"Error al obtener pacientes menores: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER PACIENTE POR ID
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['GET'])
def getPaciente(pac_id):
    """Obtiene un paciente específico por su ID"""
    pacientedao = PacienteDao()
    
    try:
        paciente = pacientedao.getPacienteById(pac_id)
        
        if paciente:
            return jsonify({
                'success': True,
                'data': paciente,
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente con el ID proporcionado.'
            }), 404
    
    except Exception as e:
        app.logger.error(f"Error al obtener el paciente: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500


# ============================================
# OBTENER PACIENTE PARA EDITAR
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>/editar', methods=['GET'])
def getPacienteParaEditar(pac_id):
    """Obtiene paciente con IDs originales para formulario de edición"""
    pacientedao = PacienteDao()

    try:
        paciente = pacientedao.getPacienteParaEditar(pac_id)

        if paciente:
            return jsonify({
                'success': True,
                'data': paciente,
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente.'
            }), 404

    except Exception as e:
        app.logger.error(f"Error al obtener paciente para editar: {str(e)}")
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno.'
        }), 500


# ============================================
# CREAR NUEVO PACIENTE
# ============================================
@pacienteapi.route('/pacientes', methods=['POST'])
def addPaciente():
    """
    Crea un nuevo paciente con todos sus datos.
    """
    data = request.get_json()
    pacientedao = PacienteDao()

    # ✅ CORRECCIÓN: Solo estos 4 campos son REALMENTE obligatorios
    campos_requeridos = ['nombre', 'apellido', 'cedula', 'fecha_nacimiento']

    # Validar campos obligatorios
    for campo in campos_requeridos:
        if campo not in data or not data[campo]:
            return jsonify({
                'success': False,
                'error': f'El campo {campo} es obligatorio y no puede estar vacío.'
            }), 400

    try:
        paciente_id = pacientedao.guardarPaciente(
            # Obligatorios
            nombre=data['nombre'],
            apellido=data['apellido'],
            cedula=data['cedula'],
            fecha_nacimiento=data['fecha_nacimiento'],
            
            # ✅ Ahora telefono también es opcional (como género/ciudad)
            telefono=data.get('telefono'),
            
            # Opcionales (ya estaban bien)
            id_genero=data.get('id_genero'),
            id_estado_civil=data.get('id_estado_civil'),
            correo=data.get('correo'),
            domicilio=data.get('domicilio'),
            id_ciudad=data.get('id_ciudad'),
            id_ciudad_nacimiento=data.get('id_ciudad_nacimiento'),
            id_nivel_instruccion=data.get('id_nivel_instruccion'),
            id_profesion=data.get('id_profesion'),
            
            # Datos de paciente
            historia_clinica=data.get('historia_clinica'),
            observaciones=data.get('observaciones'),
            
            # Datos del menor
            nom_madre=data.get('nom_madre'),
            tel_madre=data.get('tel_madre'),
            nom_padre=data.get('nom_padre'),
            tel_padre=data.get('tel_padre'),
            educacion=data.get('educacion'),
            colegio=data.get('colegio'),
            tel_colegio=data.get('tel_colegio')
        )

        if paciente_id is not None:
            return jsonify({
                'success': True,
                'data': {
                    'id_paciente': paciente_id,
                    'mensaje': 'Paciente creado exitosamente'
                },
                'error': None
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo crear el paciente. Verifique que: 1) La fecha de nacimiento sea válida, 2) Si es menor de edad, proporcione al menos el nombre de la madre o padre, 3) La historia clínica no esté duplicada.'
            }), 400

    except Exception as e:
        app.logger.error(f"Error inesperado al agregar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno del servidor. Consulte con el administrador.'
        }), 500


# ============================================
# ACTUALIZAR PACIENTE EXISTENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['PUT'])
def updatePaciente(pac_id):
    data = request.get_json()
    pacientedao = PacienteDao()

    # Verificar que existe
    paciente_existente = pacientedao.getPacienteById(pac_id)
    if not paciente_existente:
        return jsonify({
            'success': False,
            'error': 'No se encontró el paciente con el ID proporcionado.'
        }), 404

    # ✅ CORRECCIÓN: historia_clinica es obligatorio en UPDATE, pero telefono NO
    campos_requeridos = ['nombre', 'apellido', 'cedula', 'fecha_nacimiento', 'historia_clinica']

    for campo in campos_requeridos:
        if campo not in data or not data[campo]:
            return jsonify({
                'success': False,
                'error': f'El campo {campo} es obligatorio y no puede estar vacío.'
            }), 400

    try:
        resultado = pacientedao.updatePaciente(
            pac_id=pac_id,
            
            # Obligatorios
            nombre=data['nombre'],
            apellido=data['apellido'],
            cedula=data['cedula'],
            fecha_nacimiento=data['fecha_nacimiento'],
            historia_clinica=data['historia_clinica'],
            
            # ✅ Telefono ahora opcional
            telefono=data.get('telefono'),
            
            # Opcionales
            id_genero=data.get('id_genero'),
            id_estado_civil=data.get('id_estado_civil'),
            correo=data.get('correo'),
            domicilio=data.get('domicilio'),
            id_ciudad=data.get('id_ciudad'),
            id_ciudad_nacimiento=data.get('id_ciudad_nacimiento'),
            id_nivel_instruccion=data.get('id_nivel_instruccion'),
            id_profesion=data.get('id_profesion'),
            
            # Datos de paciente
            observaciones=data.get('observaciones'),
            
            # Datos del menor
            nom_madre=data.get('nom_madre'),
            tel_madre=data.get('tel_madre'),
            nom_padre=data.get('nom_padre'),
            tel_padre=data.get('tel_padre'),
            educacion=data.get('educacion'),
            colegio=data.get('colegio'),
            tel_colegio=data.get('tel_colegio')
        )

        if resultado:
            return jsonify({
                'success': True,
                'data': {
                    'id_paciente': pac_id,
                    'mensaje': 'Paciente actualizado exitosamente'
                },
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo actualizar el paciente.'
            }), 400

    except Exception as e:
        app.logger.error(f"Error al actualizar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno.'
        }), 500

# ============================================
# ELIMINAR PACIENTE
# ============================================
@pacienteapi.route('/pacientes/<int:pac_id>', methods=['DELETE'])
def deletePaciente(pac_id):
    """
    Elimina un paciente y todos sus datos asociados.
    Eliminación en cascada: pacientes_menores -> pacientes -> personas
    """
    pacientedao = PacienteDao()

    try:
        if pacientedao.deletePaciente(pac_id):
            return jsonify({
                'success': True,
                'mensaje': f'Paciente con ID {pac_id} eliminado correctamente.',
                'error': None
            }), 200
        else:
            return jsonify({
                'success': False,
                'error': 'No se encontró el paciente con el ID proporcionado o no se pudo eliminar.'
            }), 404

    except Exception as e:
        app.logger.error(f"Error al eliminar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno. Consulte con el administrador.'
        }), 500@pacienteapi.route('/pacientes', methods=['POST'])
def addPaciente():
    """
    Crea un nuevo paciente con todos sus datos.
    """
    data = request.get_json()
    pacientedao = PacienteDao()

    # ✅ CORRECCIÓN: Solo estos 4 campos son REALMENTE obligatorios
    campos_requeridos = ['nombre', 'apellido', 'cedula', 'fecha_nacimiento']

    # Validar campos obligatorios
    for campo in campos_requeridos:
        if campo not in data or not data[campo]:
            return jsonify({
                'success': False,
                'error': f'El campo {campo} es obligatorio y no puede estar vacío.'
            }), 400

    try:
        paciente_id = pacientedao.guardarPaciente(
            # Obligatorios
            nombre=data['nombre'],
            apellido=data['apellido'],
            cedula=data['cedula'],
            fecha_nacimiento=data['fecha_nacimiento'],
            
            # ✅ Ahora telefono también es opcional (como género/ciudad)
            telefono=data.get('telefono'),
            
            # Opcionales (ya estaban bien)
            id_genero=data.get('id_genero'),
            id_estado_civil=data.get('id_estado_civil'),
            correo=data.get('correo'),
            domicilio=data.get('domicilio'),
            id_ciudad=data.get('id_ciudad'),
            id_ciudad_nacimiento=data.get('id_ciudad_nacimiento'),
            id_nivel_instruccion=data.get('id_nivel_instruccion'),
            id_profesion=data.get('id_profesion'),
            
            # Datos de paciente
            historia_clinica=data.get('historia_clinica'),
            observaciones=data.get('observaciones'),
            
            # Datos del menor
            nom_madre=data.get('nom_madre'),
            tel_madre=data.get('tel_madre'),
            nom_padre=data.get('nom_padre'),
            tel_padre=data.get('tel_padre'),
            educacion=data.get('educacion'),
            colegio=data.get('colegio'),
            tel_colegio=data.get('tel_colegio')
        )

        if paciente_id is not None:
            return jsonify({
                'success': True,
                'data': {
                    'id_paciente': paciente_id,
                    'mensaje': 'Paciente creado exitosamente'
                },
                'error': None
            }), 201
        else:
            return jsonify({
                'success': False,
                'error': 'No se pudo crear el paciente. Verifique que: 1) La fecha de nacimiento sea válida, 2) Si es menor de edad, proporcione al menos el nombre de la madre o padre, 3) La historia clínica no esté duplicada.'
            }), 400

    except Exception as e:
        app.logger.error(f"Error inesperado al agregar paciente: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'error': 'Ocurrió un error interno del servidor. Consulte con el administrador.'
        }), 500